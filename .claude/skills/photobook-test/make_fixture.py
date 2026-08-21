# 사진첩 기능 테스트 데이터 생성 — 집계표 xlsx + 사진 폴더
import os, random
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook

OUT = r"D:\시스템운영-절대지우지마시오(26년)\user\Desktop\AI활용\Quickspect_테스트데이터"
PHOTO = os.path.join(OUT, "사진")
os.makedirs(PHOTO, exist_ok=True)


def make_base(seed):
    """현장사진 대용 — 세로줄 그라데이션에 균열 비슷한 선을 그려 넣는다"""
    rnd = random.Random(seed)
    img = Image.new("RGB", (900, 675))
    d = ImageDraw.Draw(img)
    tone = rnd.randint(120, 190)
    for x in range(900):
        v = tone + int(30 * (x / 900))
        d.line([(x, 0), (x, 675)], fill=(v, v - 6, v - 14))
    for _ in range(rnd.randint(2, 5)):
        x0, y0 = rnd.randint(0, 900), rnd.randint(130, 675)
        pts = [(x0, y0)]
        for _ in range(rnd.randint(4, 9)):
            x0 += rnd.randint(-70, 90); y0 += rnd.randint(-40, 60)
            pts.append((x0, y0))
        d.line(pts, fill=(70, 60, 55), width=rnd.randint(2, 5))
    return img


bases = [make_base(i) for i in range(6)]

PARTS = ["벽체", "천장 슬래브", "천장 보", "기둥", "계단 슬래브", "처마"]
DMG = ["수평균열", "수직균열", "사선균열", "망상균열", "수평 및 수직균열",
       "도장박리", "누수흔적", "균열 및 도장박리"]

# (조사위치, 건수, 파일접두)  — 실무 순서: 옥상 → 지상 → 지하 → 외부
FLOORS = [
    ("옥상층",  5, "R"),
    ("지상2층", 15, "2"),
    ("지상1층", 13, "1"),
    ("지하1층",  4, "B1"),
    ("외부",     3, "W"),
]

# 일부러 빠뜨릴 사진 (사진 없음 + 경고 팝업 확인용)
MISSING = {"205", "110", "B103"}

random.seed(7)
wb = Workbook()
ws = wb.active
ws.title = "외관집계표"

# 1~3행 머리글 (실제 파일과 같은 구조 — 데이터는 4행부터)
ws["A1"] = "외 관 조 사 결 과 표"
ws["A3"] = "번호"; ws["B3"] = "조사위치"; ws["C3"] = "세부위치"; ws["D3"] = "손상내용"
ws["E3"] = "폭(mm)"; ws["F3"] = "길이(m)"; ws["G3"] = "너비(m)"; ws["H3"] = "개소"
ws["N3"] = "비 고"; ws["P3"] = "사진첩 표기 내용"

row = 4
made = []
for loc, cnt, pfx in FLOORS:
    for n in range(1, cnt + 1):
        part = random.choice(PARTS)
        dmg = random.choice(DMG)
        w = round(random.choice([0.1, 0.2, 0.3]), 1) if "균열" in dmg else ""
        l = round(random.uniform(0.3, 3.0), 1) if w else ""
        note = "신규" if random.random() < 0.15 else ""
        # 3건 정도는 (보수완료) 파란 글씨 확인용
        done = " (보수완료)" if random.random() < 0.1 else ""

        size = f" ({w}x{l})" if w else ""
        cap = ("(신규) " if note else "") + f"{loc} {part} {dmg}{size}{done}"

        ws.cell(row, 1, n)
        ws.cell(row, 2, loc)
        ws.cell(row, 3, part)
        ws.cell(row, 4, dmg)
        if w: ws.cell(row, 5, w); ws.cell(row, 6, l)
        ws.cell(row, 8, 1)
        if note: ws.cell(row, 14, note)
        ws.cell(row, 16, cap)          # P열
        row += 1

        fname = f"{pfx}{n:02d}" if pfx in ("R", "W", "B1") else f"{pfx}{n:02d}"
        made.append((fname, f"{loc} {n}"))

wb.save(os.path.join(OUT, "외관집계표_테스트.xlsx"))

# 사진 생성 — 번호를 크게 찍어 매칭 확인이 눈으로 되게 한다
try:
    font = ImageFont.truetype("C:/Windows/Fonts/malgunbd.ttf", 96)
except Exception:
    font = ImageFont.load_default()

n_made = 0
for fname, caption in made:
    if fname in MISSING:
        continue
    src = random.choice(bases).copy()
    src = src.resize((900, 675))
    d = ImageDraw.Draw(src)
    d.rectangle([0, 0, 900, 130], fill=(0, 0, 0))
    d.text((24, 12), fname, font=font, fill=(255, 220, 0))
    src.save(os.path.join(PHOTO, fname + ".jpg"), quality=88)
    n_made += 1

print("집계표 행수:", len(made))
print("사진 생성:", n_made, "장 (일부러 빠뜨림:", sorted(MISSING), ")")
print("출력 폴더:", OUT)
